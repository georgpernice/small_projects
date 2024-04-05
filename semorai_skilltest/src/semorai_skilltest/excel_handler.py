"""Define ExcelReader class."""

from pathlib import Path
import openpyxl


class ExcelHandler:
    """Bundle methods to wrap extraction of product ID and URL."""

    def __init__(self, path_to_excel: Path, path_to_log_excels: Path) -> None:
        """Initialize ExcelReader passing path to excel.

        The Excel file must contain product IDs in col A and product URLs in col B
        In the first row shall be located only the headings."""
        self.log_folder = path_to_log_excels
        self.wb_obj = openpyxl.load_workbook(path_to_excel)
        self.sheet_obj = self.wb_obj.active

    def get_product_id(self, row_in_excel):
        """Wraps code for getting the product ID from the excel row."""

        return self.sheet_obj.cell(row=row_in_excel, column=1).value

    def get_product_url(self, row_in_excel):
        """Wraps code for getting the product URL from the excel row."""
        return self.sheet_obj.cell(row=row_in_excel, column=2).value

    def save_failed_url_product_ids(self, ids: list[int], title: str):
        """Log all IDs whose URLs failed into a new excel file."""
        wb = openpyxl.Workbook()
        cell_a1 = wb.active.cell(row=1, column=1)
        cell_a1.value = title
        for i, prod_id in enumerate(ids):
            cell = wb.active.cell(row=i + 1, column=1)
            cell.value = prod_id
        logfile = self.log_folder / "invalidURLs.xlsx"
        wb.save(logfile)

    def save_no_url_product_ids(self, ids: list[int], title: str):
        """Log all IDs with missing URLs into a new excel file."""
        wb = openpyxl.Workbook()
        cell_a1 = wb.active.cell(row=1, column=1)
        cell_a1.value = title
        for i, prod_id in enumerate(ids):
            cell = wb.active.cell(row=i + 1, column=1)
            cell.value = prod_id
        logfile = self.log_folder / "noURLs.xlsx"
        wb.save(logfile)
