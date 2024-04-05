"""Unittests for ExcelReader."""

from pathlib import Path
import shutil
import unittest

import openpyxl

from src.semorai_skilltest.excel_handler import ExcelHandler


EXCEL_PATH = Path(__file__).parent / "resources" / "Test_data.xlsx"
LOG_PATH = Path(__file__).parent / "logs"


class TestExcelHandler(unittest.TestCase):
    """Test reading methods of ExcelReader class."""

    def setUp(self) -> None:
        """Initialize an ExcelReader instance for later."""
        self.reader = ExcelHandler(EXCEL_PATH, LOG_PATH)
        LOG_PATH.mkdir(exist_ok=False)

    def tearDown(self) -> None:
        shutil.rmtree(LOG_PATH)

    def test_get_product_number(self: "TestExcelHandler"):
        """Test that method reads the product link from excel."""
        row_a2 = 2
        expected_content_a2 = 1
        cell_a2 = self.reader.get_product_id(row_a2)
        assert expected_content_a2 == cell_a2

    def test_get_product_url(self: "TestExcelHandler"):
        """Test that method reads the product URL from excel."""
        row_b2 = 2
        expected_content_b2 = (
            "https://www.murata.com/en-global/api/"
            + "pdfdownloadapi?cate=luCeramicCapacitorsSMD&partno=GRM319R71H224KA01#"
        )
        cell_b2 = self.reader.get_product_url(row_b2)
        assert expected_content_b2 == cell_b2

    def test_save_no_url_product_ids(self: "TestExcelHandler"):
        """Test that product ids are saved in a list in new excel file called 'noURL.xlsx'.

        Test that the excel is created.
        Test that excel cell a1 is correct."""

        expected = LOG_PATH / "noURLs.xlsx"

        assert expected.exists()
        wb_obj = openpyxl.load_workbook(expected)
        sheet = wb_obj.active
        a1 = sheet.cell(row=1, column=1).value
        assert a1 == "Product Numbers not associated with URLs:"

    def test_save_failed_url_product_ids(self: "TestExcelHandler"):
        """Test that product ids are saved in a list in new excel file called 'failed.xlsx'.

        Test that the excel is created.
        Test that excel contains passed product IDs."""
        expected = LOG_PATH / "invalidURLs.xlsx"

        assert expected.exists()
        wb_obj = openpyxl.load_workbook(expected)
        sheet = wb_obj.active
        a1 = sheet.cell(row=1, column=1).value
        assert a1 == "Product Numbers with invalid URLs:"
